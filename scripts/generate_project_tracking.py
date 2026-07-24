import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "203764"   # Dark blue
C_HEADER_FG   = "FFFFFF"   # White
C_ALT_ROW     = "F2F2F2"   # Light grey
C_WHITE       = "FFFFFF"
C_BORDER      = "A6A6A6"   # Grey border
C_INSTRUCTION_BG = "FFF2CC" # Light yellow for instruction

def make_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project"
    ws.sheet_view.showGridLines = False

    # Row 1: Title
    ws.merge_cells("A1:J1")
    c1 = ws["A1"]
    c1.value = "Total Project Tracking - Parking Building Management System (PBMS)"
    c1.font = Font(name="Arial", bold=True, size=16, color="000000")
    c1.alignment = center()
    
    # Row 2: Instructions
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = "Information in the columns A-E are filled in the project initiation; columns F-J to be filled by the end of each development iteration"
    c2.font = Font(name="Arial", italic=True, size=10, color="000000")
    c2.fill = fill(C_INSTRUCTION_BG)
    c2.alignment = left()

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 30

    # Row 3: Headers
    headers = [
        "#", 
        "Screen/Function", 
        "Feature", 
        "Actor", 
        "Screen/Function Description", 
        "In Charge", 
        "Status", 
        "Actual", 
        "Updated", 
        "Update Details"
    ]
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=11, color=C_HEADER_FG)
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = make_border()

    # Data rows
    # Map roles
    # Nguyễn Duy Hiếu: PM, Backend
    # Nguyễn Mạnh Hưng: Backend, Documentation
    # Nguyễn Quang Huy: UI, Prototype, Frontend
    # Trần Thái Bình: Documentation, UI, PaddleOCR

    tasks = [
        # Phase 1: Initiation & Design
        ("Initiation", "Project Scoping", "System", "Define PBMS requirements, hardware limits, and core features", "Nguyễn Duy Hiếu", "Done", "Sprint 1", "01-Jun-2026", "Finalised project scope and DB schema"),
        ("Design", "UI/UX Prototyping - Driver", "Driver", "Figma prototype for Driver mobile web portal", "Nguyễn Quang Huy", "Done", "Sprint 1", "05-Jun-2026", "Completed wireframes and color scheme"),
        ("Design", "UI/UX Prototyping - Staff/Manager", "Staff/Manager", "Figma prototype for desktop monitoring dashboard", "Trần Thái Bình", "Done", "Sprint 1", "08-Jun-2026", "Completed layout of slot map and KPI charts"),
        ("Design", "Architecture & DB Setup", "System", "Setup PostgreSQL schema with proper indexes, constraints and isolation levels", "Nguyễn Duy Hiếu", "Done", "Sprint 1", "10-Jun-2026", "Used EF Core Code-First Migrations"),
        ("Design", "Technical Documentation (SRS/SDS)", "System", "Drafting Software Requirements and Software Design Specifications", "Nguyễn Mạnh Hưng", "Done", "Sprint 1", "12-Jun-2026", "Completed system topologies"),
        ("DevOps", "CI/CD Deployment", "System", "Auto build/test on Github Actions; deploy backend to Railway and Frontend to Firebase", "Nguyễn Duy Hiếu", "Done", "Sprint 1", "14-Jun-2026", "Implemented workflow yaml with Gitleaks scan"),

        # Phase 2: Core Platform & Auth
        ("Authentication", "Login (JWT)", "All Roles", "User attempts to log in via JWT token (email/password)", "Nguyễn Mạnh Hưng", "Done", "Sprint 2", "16-Jun-2026", "Implemented BCrypt and JWT provider"),
        ("Authentication", "Google Auth SSO", "Driver", "Driver logs in / registers using Google OAuth Single Sign-on", "Nguyễn Mạnh Hưng", "Done", "Sprint 2", "18-Jun-2026", "Migrated to Google.Apis.Auth validation"),
        ("Frontend", "Project Scaffold & UI Library", "System", "Setup Vite React, TailwindCSS, Router and Axios handlers", "Nguyễn Quang Huy", "Done", "Sprint 2", "19-Jun-2026", "Configured layout and component library"),
        
        # Phase 3: Driver Module Core
        ("Driver Module", "Current Session UI", "Driver", "Live tile showing current parking session, location (floor/slot), entry time", "Nguyễn Quang Huy", "Done", "Sprint 3", "22-Jun-2026", "Used React interval to refresh fee simulation"),
        ("Driver Module", "Wallet Top-Up UI", "Driver", "Integration of PayOS QR generation and top-up UI workflow", "Nguyễn Quang Huy", "Done", "Sprint 3", "24-Jun-2026", "Implemented polling for payment success status"),
        ("Driver Backend", "Wallet PayOS Webhook", "System", "Handle PayOS webhooks for payment success verification", "Nguyễn Duy Hiếu", "Done", "Sprint 3", "25-Jun-2026", "Idempotency handling and balance updates"),
        ("Driver Module", "Active Bookings UI", "Driver", "User interface for reserving parking slots ahead of time", "Nguyễn Quang Huy", "Done", "Sprint 4", "26-Jun-2026", "Handled double-booking bug on frontend"),
        ("Driver Backend", "Booking Engine APIs", "System", "API endpoints and background expiry service for reservations", "Nguyễn Mạnh Hưng", "Done", "Sprint 4", "28-Jun-2026", "Reservation expiration CRON job created"),
        ("Driver Module", "Monthly Pass Subscriptions", "Driver", "Subscription purchase and viewing interface", "Trần Thái Bình", "Done", "Sprint 5", "05-Jul-2026", "Frontend design for monthly passes"),
        ("Driver Backend", "Subscription API & Pricing", "System", "Pricing calculation bypass for subscribers, billing API", "Nguyễn Duy Hiếu", "Done", "Sprint 5", "10-Jul-2026", "Added subscription waiver logic"),
        
        # Phase 4: Staff Operations & OCR
        ("Staff Module", "Slot Map Visualiser UI", "Staff", "Real-time slot grid with cell colors (Occupied/Available)", "Nguyễn Quang Huy", "Done", "Sprint 3", "20-Jun-2026", "Figma Design implemented in React"),
        ("Staff Backend", "SignalR Real-Time Comm", "System", "Setup SignalR Hub to broadcast slot activity", "Nguyễn Mạnh Hưng", "Done", "Sprint 3", "22-Jun-2026", "Created broadcast groups logic (all, operations)"),
        ("Staff Backend", "Check-In Processing", "Staff", "Atomic transaction for check-in: slot assigned, session initiated", "Nguyễn Duy Hiếu", "Done", "Sprint 3", "24-Jun-2026", "Uses Serializable transactions. Prevents duplicates."),
        ("Staff Backend", "Check-Out Processing", "Staff", "Calculate final fee based on duration/grace period; release slot", "Nguyễn Duy Hiếu", "Done", "Sprint 3", "25-Jun-2026", "Integrated with fee calculator rules (e.g. lost ticket)"),
        ("Staff Module", "Camera OCR Integration", "Staff", "PaddleOCR server implementation for license plate scanning", "Trần Thái Bình", "Done", "Sprint 4", "10-Jul-2026", "Python PaddleOCR wrapper + PlateRecognizer fallback fallback"),
        ("Staff Module", "Check-in Booking Auto-Redeem", "Staff", "UI modifications to auto-fetch booking and Backend API support", "Nguyễn Mạnh Hưng", "Done", "Sprint 5", "24-Jul-2026", "Auto mapped by frontend calling by-plate API before check-in"),
        ("Staff Module", "Incident Logging UI/UX", "Staff", "Form for staff to create and track rule violations (e.g. lost ticket)", "Trần Thái Bình", "Done", "Sprint 5", "15-Jul-2026", "Assisted in frontend binding"),

        # Phase 5: Administration & Delivery
        ("Manager Module", "Dashboard KPIs Setup", "Manager", "Overall building usage, daily revenue sums, number of active sessions", "Nguyễn Mạnh Hưng", "Done", "Sprint 5", "16-Jul-2026", "Implemented snapshot DB queries"),
        ("Manager Module", "Dashboard UI Charts", "Manager", "Implemented Recharts for 30-day analytics and floor density", "Nguyễn Quang Huy", "Done", "Sprint 5", "18-Jul-2026", "React-based interactive components"),
        ("Admin Module", "Zone & Slot Setup", "Admin", "Define floors, zones, vehicle type constraints, base price policies", "Nguyễn Duy Hiếu", "Done", "Sprint 1", "30-May-2026", "CRUD ops provided via DatabaseSeeder base"),
        ("Security", "Rate Limiting & CORS", "System", "Defend auth endpoints against brute force; lockdown origin access", "Nguyễn Duy Hiếu", "Done", "Sprint 4", "10-Jul-2026", "Applied standard .NET middlewares"),
        ("Documentation", "Testing Report & UML", "System", "Full end-to-end testing, UML refactoring and Test cases generation", "Trần Thái Bình", "Done", "Sprint 5", "20-Jul-2026", "Verified UI tests and PaddleOCR stability"),
        ("Documentation", "User Manual & Presentations", "System", "Finalising slides and operation instructions for the evaluation board", "Nguyễn Mạnh Hưng", "Done", "Sprint 5", "23-Jul-2026", "Assisted by Trần Thái Bình"),
    ]

    r = 4
    for idx, row_data in enumerate(tasks, 1):
        bg = C_ALT_ROW if idx % 2 == 0 else C_WHITE
        
        # "#" Column
        ws.cell(row=r, column=1, value=idx).alignment = center()
        
        # Write remaining
        col = 2
        for val in row_data:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(bg)
            c.font = Font(name="Arial", size=10)
            c.border = make_border()
            
            # Alignments
            if col in [7, 8, 9]:
                c.alignment = center()
            else:
                c.alignment = left()
            
            col += 1
            
        # Re-apply border and style to column 1
        c_1 = ws.cell(row=r, column=1)
        c_1.fill = fill(bg)
        c_1.border = make_border()
        c_1.font = Font(name="Arial", size=10)
        
        ws.row_dimensions[r].height = 40
        r += 1

    # Adjust widths
    widths = [5, 18, 25, 12, 50, 15, 12, 12, 12, 45]
    set_col_widths(ws, widths)
    
    # Save Output
    out_dir = r"f:\FPT_material\2026\Summer\SWP391\Project\Parking-Building-Management-System\docs\Exported_Plain"
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, "Project_Tracking_PBMS_SU26SWP08.xlsx")
    wb.save(out_file)
    print(f"Generated {out_file}")

if __name__ == "__main__":
    main()
