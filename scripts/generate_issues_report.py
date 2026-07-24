import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SUB_BG      = "2E75B6"   # medium blue
C_SUB_FG      = "FFFFFF"
C_ALT_ROW     = "EBF5FB"   # very light blue
C_WHITE       = "FFFFFF"
C_BORDER      = "2E75B6"

def make_border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_header_row(ws, row, values, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = make_border()

def apply_body_row(ws, row, values, alt=False, bold_first=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=(bold_first and col == 1), size=10, color="1F1F1F")
        c.fill = fill(bg)
        c.alignment = center() if col in [1, 4, 5] else left()
        c.border = make_border("AABCD4", "thin")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Cover & Issue Register
    # ═══════════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "1. Issue Register"
    ws1.sheet_view.showGridLines = False

    # Cover Header
    ws1.merge_cells("A2:E4")
    tc = ws1["A2"]
    tc.value = "ISSUES REPORT"
    tc.font = Font(name="Calibri", bold=True, size=28, color=C_HEADER_FG)
    tc.fill = fill(C_HEADER_BG)
    tc.alignment = center()

    ws1.merge_cells("A5:E5")
    sub = ws1["A5"]
    sub.value = "Parking Building Management System (PBMS) — SU26SWP08"
    sub.font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    sub.fill = fill(C_SUB_BG)
    sub.alignment = center()

    ws1.row_dimensions[2].height = 30
    ws1.row_dimensions[3].height = 30
    ws1.row_dimensions[4].height = 30
    ws1.row_dimensions[5].height = 25

    # Document Info
    info = [
        ("Project Code", "SU26SWP08"),
        ("Document Type", "Issues Report (Template 4)"),
        ("Version", "1.1"),
        ("Period", "Sprint 1 – Sprint Final (June – July 2026)"),
        ("Prepared By", "SU26SWP08 Development Team")
    ]
    r = 7
    for i, (key, val) in enumerate(info):
        bg = C_ALT_ROW if i % 2 == 0 else C_WHITE
        ws1.merge_cells(f"A{r}:B{r}")
        c1 = ws1[f"A{r}"]
        c1.value = key
        c1.font = Font(name="Calibri", bold=True, size=10)
        c1.fill = fill(bg)
        c1.alignment = left()
        c1.border = make_border("AABCD4")
        ws1.merge_cells(f"C{r}:E{r}")
        c2 = ws1[f"C{r}"]
        c2.value = val
        c2.font = Font(name="Calibri", size=10)
        c2.fill = fill(bg)
        c2.alignment = left()
        c2.border = make_border("AABCD4")
        r += 1

    r += 2
    ws1.merge_cells(f"A{r}:E{r}")
    ws1[f"A{r}"].value = "1. Issue Register"
    ws1[f"A{r}"].font = Font(name="Calibri", bold=True, size=12, color=C_SUB_FG)
    ws1[f"A{r}"].fill = fill(C_SUB_BG)
    ws1[f"A{r}"].alignment = left(wrap=False)
    ws1[f"A{r}"].border = make_border()
    r += 1

    headers = ["Issue ID", "Title", "Category", "Severity", "Status"]
    apply_header_row(ws1, r, headers)
    r += 1

    issues = [
        ("ISS-001", "Concurrent check-in double slot assignment", "Transaction", "Critical", "Resolved"),
        ("ISS-002", "PaddleOCR exception on Linux target", "Integration", "High", "Resolved"),
        ("ISS-003", "PayOS signature mismatch", "Integration", "High", "Resolved"),
        ("ISS-004", "Google Auth inactive on production hosting", "Configuration", "High", "Resolved"),
        ("ISS-005", "SignalR broadcast failure to Driver client", "Realtime", "Medium", "Resolved"),
        ("ISS-006", "Reservation auto-expiry scheduler error", "Background", "Medium", "Resolved"),
        ("ISS-007", "Zero fee charged for exactly 1-hour session", "Logic", "Medium", "Resolved"),
        ("ISS-008", "EWallet double deduction condition", "Transaction", "Critical", "Resolved"),
        ("ISS-009", "Cross-Origin (CORS) policy block", "Configuration", "High", "Resolved"),
        ("ISS-010", "Manager dashboard timestamp offset", "UI", "Low", "Resolved"),
        ("ISS-011", "Ticket lookup intuitively triggered checkout", "UX/Logic", "Medium", "Resolved"),
        ("ISS-012", "Slot map checkout bypassed confirmation", "UX/Logic", "High", "Resolved"),
        ("ISS-013", "Double booking on spam click", "Transaction", "Medium", "Resolved"),
        ("ISS-014", "Check-in did not auto-redeem reservation", "Logic", "High", "Resolved"),
    ]

    for i, row_data in enumerate(issues):
        alt = (i % 2 == 1)
        apply_body_row(ws1, r, row_data, alt=alt, bold_first=True)
        ws1.row_dimensions[r].height = 25
        r += 1

    set_col_widths(ws1, [15, 60, 20, 15, 15])

    # ═══════════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Resolution Details
    # ═══════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("2. Issue Resolution Details")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "2. Issue Resolution Details"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws2["A1"].fill = fill(C_HEADER_BG)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    headers2 = ["Issue ID", "Title", "Description", "Root Cause", "Resolution", "Affected Components", "Status"]
    apply_header_row(ws2, 2, headers2)

    details = [
        ("ISS-001", "Concurrent check-in double slot assignment",
         "Simultaneous check-in requests generated duplicate Active session records for a single slot ID over separate threads.",
         "Database read/write isolated under lower isolation bounds, allowing dirty reads.",
         "Implementation of explicit ExecuteInTransactionAsync(IsolationLevel.Serializable) around the check-in data context scope.",
         "ParkingSessionService.cs", "Resolved"),
        
        ("ISS-002", "PaddleOCR exception on Linux target",
         "Platform threw DllNotFoundException when deployed to Linux nodes.",
         "PaddleInference requires native Windows .dll build definitions not present in target environment.",
         "Bound implementation to OperatingSystem.IsWindows(); injected stub service for Linux runtimes. Added web-based fallback endpoint parameters.",
         "LocalOcrEngine.cs", "Resolved"),

        ("ISS-003", "PayOS Webhook Checksum Validation",
         "Incoming webhooks systematically failed HMAC validation routines.",
         "Payload object properties serialized to string incorrectly (order mismatch).",
         "Implemented custom payload flattening to sort keys alphabetically as required by PayOS security standards prior to computation.",
         "PayOsWebhookController.cs", "Resolved"),

        ("ISS-005", "SignalR Hub Group Assignment",
         "Slot map failed to trigger state updates on the Driver's viewport automatically.",
         "Driver connections lacked inclusion in the correct target broadcast group mapping.",
         "Appended explicit Groups.AddToGroupAsync assignment in the Hub configuration for authenticated standard users.",
         "ParkingHub.cs", "Resolved"),

        ("ISS-011", "Ticket lookup intuitively triggered checkout",
         "Searching for a ticket code immediately checked the car out without confirming the fee.",
         "The handleSearchTicket handler invoked the checkout API via processPlate directly.",
         "Refactored to fetch session details and open a Checkout Confirmation Modal to review fees and payment methods.",
         "StaffDashboard.tsx", "Resolved"),
        
        ("ISS-012", "Slot map checkout bypassed confirmation",
         "Clicking on an occupied slot in the dashboard immediately checked the car out.",
         "onClick handler triggered checkout instead of merely selecting the slot.",
         "Removed auto-checkout on slot click. Added checkout flow via a detail panel and a confirmation modal.",
         "StaffDashboard.tsx, ParkingSlotMap.tsx", "Resolved"),

        ("ISS-013", "Double booking on spam click",
         "Driver users could book duplicate reservations for the same license plate if they spam clicked.",
         "Reservation UI lacked a submit guard and backend lacked a unique plate check.",
         "Added useRef-based submit guards on Frontend and added active reservation limits/duplicate checks on Backend API.",
         "UserMobileHome.tsx, ReservationService.cs", "Resolved"),

        ("ISS-014", "Check-in did not auto-redeem reservation",
         "When a reserved vehicle arrived, scanning the plate would create a new standard checkin, ignoring the booking.",
         "Checkin flow did not cross-reference pending reservations for the scanned plate.",
         "Created a by-plate reservation lookup API and automated the match in the staff UI during check-in.",
         "StaffDashboard.tsx, ReservationsController.cs", "Resolved"),
    ]

    r = 3
    for i, row_data in enumerate(details):
        alt = (i % 2 == 1)
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = Font(name="Calibri", bold=(col == 1), size=10, color="1F1F1F")
            c.fill = fill(C_ALT_ROW if alt else C_WHITE)
            c.alignment = Alignment(horizontal="center" if col in [1,7] else "left", vertical="top", wrap_text=True)
            c.border = make_border("AABCD4", "thin")
        ws2.row_dimensions[r].height = 80
        r += 1

    set_col_widths(ws2, [15, 30, 45, 40, 45, 25, 12])
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Discovered Limitations
    # ═══════════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("3. Discovered Limitations")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    ws3["A1"].value = "3. Discovered Limitations"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws3["A1"].fill = fill(C_HEADER_BG)
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    apply_header_row(ws3, 2, ["Limitation ID", "Component", "Description"], size=10)

    limitations = [
        ("LIM-001", "OCR Engine", "Initial PaddleOCR library initialization blocks the main thread for 8-12 seconds on the primary call. A background warmup sequence was deployed to mitigate latency post-startup."),
        ("LIM-002", "Report Generation", "Execution of snapshot generation requires manual administrator trigger due to omission of automated CRON scheduling logic in standard builds."),
        ("LIM-003", "Email Notification", "Currently SMTP emailing for subscription expiry uses a synchronous task block. Recommended moving to background Queue channel."),
    ]

    r = 3
    for i, row_data in enumerate(limitations):
        alt = (i % 2 == 1)
        for col, val in enumerate(row_data, 1):
            c = ws3.cell(row=r, column=col, value=val)
            c.font = Font(name="Calibri", bold=(col == 1), size=10, color="1F1F1F")
            c.fill = fill(C_ALT_ROW if alt else C_WHITE)
            c.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="top", wrap_text=True)
            c.border = make_border("AABCD4", "thin")
        ws3.row_dimensions[r].height = 60
        r += 1

    set_col_widths(ws3, [15, 25, 80])

    # Save
    out_dir = r"f:\FPT_material\2026\Summer\SWP391\Project\Parking-Building-Management-System\docs\Exported_Plain"
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, "Issues_Report_PBMS_SU26SWP08.xlsx")
    wb.save(out_file)
    print(f"Generated {out_file}")

if __name__ == "__main__":
    main()
