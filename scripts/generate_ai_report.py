"""
Generate AI Usage Report Excel file for PBMS project (SU26SWP08)
Run: python scripts/generate_ai_report.py
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SUB_BG      = "2E75B6"   # medium blue
C_SUB_FG      = "FFFFFF"
C_SEC_BG      = "D6E4F0"   # light blue tint
C_ALT_ROW     = "EBF5FB"   # very light blue
C_WHITE       = "FFFFFF"
C_ACCENT1     = "E74C3C"   # red accent
C_ACCENT2     = "27AE60"   # green accent
C_ACCENT3     = "F39C12"   # orange accent
C_BORDER      = "2E75B6"

def make_border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(bold=True, size=11, color=C_HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="1F1F1F"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_header_row(ws, row, values, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = make_border()

def apply_body_row(ws, row, values, alt=False, bold_first=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=(bold_first and col == 1), size=10, color="1F1F1F")
        c.fill = fill(bg)
        c.alignment = left()
        c.border = make_border("AABCD4", "thin")

def section_title(ws, row, col, text, span, bg=C_SUB_BG):
    ws.cell(row=row, column=col, value=text).font = Font(name="Calibri", bold=True, size=12, color=C_SUB_FG)
    ws.cell(row=row, column=col).fill = fill(bg)
    ws.cell(row=row, column=col).alignment = left(wrap=False)
    ws.cell(row=row, column=col).border = make_border()
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Cover / Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "1. Cover & Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 10

    # Title block
    ws.merge_cells("A2:G4")
    tc = ws["A2"]
    tc.value = "AI USAGE REPORT"
    tc.font = Font(name="Calibri", bold=True, size=28, color=C_HEADER_FG)
    tc.fill = fill(C_HEADER_BG)
    tc.alignment = center()

    ws.merge_cells("A5:G5")
    sub = ws["A5"]
    sub.value = "Parking Building Management System (PBMS) — SU26SWP08"
    sub.font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    sub.fill = fill(C_SUB_BG)
    sub.alignment = center()

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 25

    # Document info table
    info = [
        ("Project Code",       "SU26SWP08"),
        ("Project Name",       "Parking Building Management System"),
        ("Document Type",      "AI Usage Report"),
        ("Document Version",   "1.0"),
        ("Report Period",      "Sprint 1 – Sprint Final  (June 2026 – July 2026)"),
        ("Prepared By",        "SU26SWP08 Development Team"),
        ("Date",               "July 2026"),
        ("Status",             "Final"),
    ]
    r = 7
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"A{r}"].value = "Document Information"
    ws[f"A{r}"].font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    ws[f"A{r}"].fill = fill(C_HEADER_BG)
    ws[f"A{r}"].alignment = center()
    ws.merge_cells(f"C{r}:G{r}")
    ws[f"C{r}"].value = "Details"
    ws[f"C{r}"].font = Font(name="Calibri", bold=True, size=11, color=C_HEADER_FG)
    ws[f"C{r}"].fill = fill(C_HEADER_BG)
    ws[f"C{r}"].alignment = center()
    r += 1

    for i, (key, val) in enumerate(info):
        bg = C_ALT_ROW if i % 2 == 0 else C_WHITE
        ws.merge_cells(f"A{r}:B{r}")
        c1 = ws[f"A{r}"]
        c1.value = key
        c1.font = Font(name="Calibri", bold=True, size=10)
        c1.fill = fill(bg)
        c1.alignment = left()
        c1.border = make_border("AABCD4")
        ws.merge_cells(f"C{r}:G{r}")
        c2 = ws[f"C{r}"]
        c2.value = val
        c2.font = body_font()
        c2.fill = fill(bg)
        c2.alignment = left()
        c2.border = make_border("AABCD4")
        r += 1

    # Executive Summary
    r += 1
    section_title(ws, r, 1, "Executive Summary", 7)
    r += 1
    summary_text = (
        "This report documents all AI tool usage during the development of the Parking Building Management System "
        "(PBMS) throughout the Summer 2026 semester. AI tools — specifically Cursor IDE (AI-powered editor), "
        "ChatGPT (GPT-4o), and GitHub Copilot — were integrated to accelerate development velocity, reduce "
        "boilerplate burden, and assist with technical documentation.\n\n"
        "All AI-generated outputs (code, document drafts, architecture commentary) were subject to mandatory "
        "developer review before integration into the codebase or documentation repository. No sensitive data "
        "(connection strings, JWT keys, PayOS secrets) was included in any AI prompt. The estimated AI drafting "
        "ratio across the codebase averages 30–40%, with 100% human review applied to every artifact."
    )
    ws.merge_cells(f"A{r}:G{r+4}")
    sc = ws[f"A{r}"]
    sc.value = summary_text
    sc.font = body_font(size=10)
    sc.fill = fill(C_WHITE)
    sc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    sc.border = make_border("AABCD4")
    ws.row_dimensions[r].height = 90

    set_col_widths(ws, [18, 18, 25, 20, 18, 18, 18])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – AI Tools Inventory
# ═══════════════════════════════════════════════════════════════════════════════
def build_tools(wb):
    ws = wb.create_sheet("2. AI Tools Inventory")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "AI Tools Inventory — All Tools Used in PBMS Development"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Tool Name", "Version / Model", "Access Method", "Primary Usage", "Frequency", "Notes"]
    apply_header_row(ws, 2, headers)

    tools = [
        ("Cursor IDE",
         "Cursor v0.40+ (GPT-4o / Claude 3.5 Sonnet backend)",
         "Licensed IDE – local installation",
         "AI-assisted code editing, inline completions, multi-file refactoring, context-aware suggestions within the .NET / React codebase",
         "Daily – entire development lifecycle",
         "Used as primary development environment. Tab-completion and Chat panel both used. No secrets shared."),
        ("ChatGPT (GPT-4o)",
         "GPT-4o (OpenAI)",
         "Web app (chat.openai.com)",
         "Architecture design consultation, code snippet generation (PayOS webhook HMAC, Serializable transaction patterns), document drafting, and SRS/SDS review",
         "Several times per sprint",
         "Used for architectural guidance and documentation. Outputs reviewed and adapted before use."),
        ("GitHub Copilot",
         "Copilot (GPT-4o Turbo)",
         "VS Code extension + JetBrains plugin",
         "Inline code suggestion for repetitive patterns: DTO property declarations, EF Core LINQ queries, xUnit test method scaffolding",
         "Daily – during active coding",
         "Suggestions accepted selectively after review. Rejected hallucinated API usage."),
        ("Antigravity / Google Gemini",
         "Gemini 2.5 Pro / Claude Sonnet 4.6",
         "Gemini workspace tool",
         "Document generation: SRS, SDS, UML diagrams, Testing Report, User Manual, AI Usage Report; review of architecture artifacts",
         "Weekly – documentation phases",
         "Primarily used for documentation writing and codebase analysis. All outputs verified against source code."),
    ]

    for i, row_data in enumerate(tools):
        alt = (i % 2 == 1)
        apply_body_row(ws, i + 3, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[i + 3].height = 65

    set_col_widths(ws, [22, 30, 28, 50, 28, 50])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Detailed Usage Log
# ═══════════════════════════════════════════════════════════════════════════════
def build_usage_log(wb):
    ws = wb.create_sheet("3. Detailed Usage Log")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Detailed AI Usage Log — Per Task / Feature"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = [
        "Log ID", "Sprint / Date", "Feature / Module", "AI Tool",
        "Prompt Category", "AI Contribution Description",
        "Developer Review / Modification", "Files Affected", "Outcome"
    ]
    apply_header_row(ws, 2, headers, size=10)
    ws.row_dimensions[2].height = 35

    # ── Rows ──────────────────────────────────────────────────────────────────
    rows = [
        # Sprint 1 – Initial Setup
        ("AI-001", "Sprint 1\n2026-05-30", "Project Scaffold\n& Monorepo Setup",
         "ChatGPT", "Architecture",
         "Consulted GPT-4o for monorepo layout (ASP.NET Core API + Vite React frontend). Suggested separation of backend/ and frontend/ with shared .gitignore, Dockerfile patterns, and GitHub Actions workflow skeleton.",
         "Reviewed recommended structure. Adapted Dockerfile to multi-stage build. Added Railway-specific toml configuration manually. Customised .gitignore to exclude appsettings.Local.json.",
         ".gitignore, Dockerfile, railway.toml, .github/workflows/",
         "Accepted with modifications"),

        ("AI-002", "Sprint 1\n2026-05-30", "CI / CD Pipeline",
         "ChatGPT", "DevOps / IaC",
         "Requested GitHub Actions workflow for .NET build, test, and Firebase deploy. Also requested Gitleaks secret-scanning integration step.",
         "Validated each job step. Corrected Firebase deploy target (hosting not functions). Added conditional CD on main branch only.",
         ".github/workflows/ci.yml, .github/workflows/cd.yml",
         "Accepted with modifications"),

        ("AI-003", "Sprint 1\n2026-05-30", "Database Schema\n(PostgreSQL)",
         "Cursor (Chat)", "Code Generation",
         "Used Cursor Chat to review initial EF Core model classes and suggest appropriate index strategies for ParkingSessions and ParkingSlots tables (composite index on LicensePlate + Status).",
         "Manually verified index choices against expected query patterns. Added missing cascade delete constraints and adjusted column nullability.",
         "Models/*.cs, Migrations/*",
         "Accepted with modifications"),

        # Sprint 2
        ("AI-004", "Sprint 2\n2026-06-16", "Authentication\n(JWT + BCrypt)",
         "Cursor (Inline)", "Code Generation",
         "Inline completions for JWT token generation logic (IssuerSigningKey setup, claim mapping) and BCrypt password hashing calls in AuthService.",
         "Reviewed token expiry (set to 24h per spec). Verified BCrypt cost factor = 11. Added role claim injection for multi-role enforcement.",
         "Services/AuthService.cs, Controllers/AuthController.cs",
         "Accepted with modifications"),

        ("AI-005", "Sprint 2\n2026-06-16", "Role-Based\nPortal Layout (React)",
         "Cursor (Chat)", "Code Generation",
         "Scaffolded React Router v6 nested route structure separating /admin, /manager, /staff, /driver portals. Generated initial ProtectedRoute component and layout wrappers.",
         "Corrected component import paths. Added token-expiry redirect logic. Integrated Axios interceptor for 401 auto-logout.",
         "frontend/src/app/router.tsx, frontend/src/app/layouts/",
         "Accepted with modifications"),

        ("AI-006", "Sprint 2\n2026-06-16", "PayOS Payment\nWebhook Handler",
         "ChatGPT", "Security / Integration",
         "Asked GPT-4o for HMAC-SHA256 signature verification pattern for incoming PayOS webhooks. Provided C# implementation template.",
         "Verified HMAC key usage matches PayOS documentation. Integrated with WalletService to credit balance atomically. Added idempotency check to prevent double-credit.",
         "Controllers/PayOsWebhookController.cs, Services/PayOs/",
         "Accepted with modifications"),

        # Sprint 3
        ("AI-007", "Sprint 3\n2026-06-20", "Parking Check-In\nWorkflow",
         "Cursor (Chat)", "Code Generation",
         "Used Cursor Chat to discuss Serializable isolation approach for the check-in transaction (slot availability + session creation + slot status update in one atomic block). Generated initial ExecuteInTransactionAsync wrapper pattern.",
         "Reviewed isolation level implications. Verified deadlock handling. Added OCR license plate pre-processing (ToUpperInvariant, Trim). Added ReservationId linking logic.",
         "Services/ParkingSessionService.cs, Data/ApplicationDbContext.cs",
         "Accepted with modifications"),

        ("AI-008", "Sprint 3\n2026-06-20", "Slot Allocation\nAlgorithm",
         "Cursor (Inline)", "Code Generation",
         "Inline suggestions for LINQ query to select least-occupied matching zone and first available slot. Used Cursor tab-completion for EF Core .Include() chains.",
         "Validated slot ordering logic (SlotPosition ascending). Added VIP slot (position 1) priority path. Tested under concurrent load.",
         "Services/SlotAllocationService.cs",
         "Accepted with modifications"),

        ("AI-009", "Sprint 3\n2026-06-20", "Pricing Service\n(Fee Calculation)",
         "GitHub Copilot", "Code Generation",
         "Copilot suggested grace-period boundary check, ceil(hours) multiplication pattern, and daily cap application in PricingService.",
         "Verified against SRS pricing formula. Corrected edge case: duration exactly = grace_period should yield zero fee. Added lost ticket fee branch and subscription waiver.",
         "Services/PricingService.cs",
         "Accepted with modifications"),

        ("AI-010", "Sprint 3\n2026-06-26", "SignalR Real-Time\nHub & Notifications",
         "Cursor (Chat)", "Code Generation",
         "Consulted Cursor Chat for SignalR group-based notification pattern (all, operations, driver:{id}). Generated hub configuration and client-side useSignalR hook scaffold.",
         "Validated hub group naming. Corrected connection lifecycle management (disconnect on unmount). Integrated with slot map component for live colour updates.",
         "Hubs/ParkingHub.cs, Services/ParkingRealtimeNotifier.cs, frontend/src/app/hooks/useSignalR.ts",
         "Accepted with modifications"),

        # Sprint 4
        ("AI-011", "Sprint 4\n2026-06-26", "Reservation\nManagement",
         "ChatGPT", "Architecture",
         "Requested GPT-4o review of reservation lifecycle (Pending → Confirmed → CheckedIn / Expired / Cancelled) and background expiry service approach using IHostedService.",
         "Validated state transitions against business rules. Implemented ReservationExpiryHostedService with configurable polling interval. Tested auto-expiry under load.",
         "Services/ReservationService.cs, Services/ReservationExpiryHostedService.cs",
         "Accepted with modifications"),

        ("AI-012", "Sprint 4\n2026-06-26", "Google OAuth\nIntegration (Drivers)",
         "Cursor (Chat)", "Integration",
         "Used Cursor Chat to scaffold Google token validation flow (ID token → GoogleTokenValidator → user upsert). Generated initial HttpClient call to Google tokeninfo endpoint.",
         "Replaced polling with Google .NET Client Library for proper verification. Added first-time registration path. Restricted OAuth to Driver role only per FR-AUTH-004.",
         "Services/GoogleTokenValidator.cs, Controllers/AuthController.cs",
         "Accepted with modifications"),

        ("AI-013", "Sprint 4\n2026-07-10", "OCR License Plate\nScanner",
         "ChatGPT", "Integration",
         "Asked GPT-4o for approach to integrate PaddleOCR (local Python server) with .NET API. Generated HttpClient wrapper and fallback to Plate Recognizer API.",
         "Implemented two-path OCR: local PaddleOCR (Windows/Linux stub) and Plate Recognizer cloud API fallback. Added Vietnamese plate regex validation post-OCR.",
         "Services/Ocr/, Controllers/PlateOcrController.cs",
         "Accepted with modifications"),

        ("AI-014", "Sprint 4\n2026-07-10", "Security Hardening\n(Rate Limiting, Headers)",
         "ChatGPT", "Security",
         "Requested rate-limiting middleware pattern for ASP.NET Core (8+ req/min on auth endpoints per FR-AUTH-006) and security headers (X-Frame-Options, X-Content-Type-Options).",
         "Integrated .NET built-in rate limiter (RateLimiterMiddleware). Verified header values. Added CORS policy restricted to Firebase frontend origin.",
         "Program.cs, Middleware/",
         "Accepted with modifications"),

        # Sprint 5
        ("AI-015", "Sprint 5\n2026-07-14", "Monthly Subscription\n(Plan + Purchase + Checkout)",
         "Cursor (Chat)", "Code Generation",
         "Used Cursor Chat to scaffold SubscriptionPlanService (CRUD), SubscriptionService (purchase with wallet deduction), and subscription-waiver check at checkout.",
         "Reviewed deduction atomicity (Serializable). Added subscription expiry background service. Verified waiver logic: active subscription → parking fee = 0 but penalty/VIP surcharge still applies.",
         "Services/SubscriptionService.cs, Services/SubscriptionPlanService.cs, Services/SubscriptionExpiryHostedService.cs",
         "Accepted with modifications"),

        ("AI-016", "Sprint 5\n2026-07-14", "Reporting Dashboard\n(KPIs + Charts)",
         "Cursor (Inline)", "Code Generation",
         "Inline completions for ReportsController LINQ aggregate queries (sum revenue, count sessions, occupancy % per floor). Suggested ReportSnapshot daily cron approach.",
         "Validated query performance. Added composite index on ParkingSessions.EntryTime. Verified revenue drill-down against payment records.",
         "Controllers/ReportsController.cs, Services/ReportSnapshotService.cs",
         "Accepted with modifications"),

        # Documentation phase
        ("AI-017", "Doc Phase\n2026-07-20", "SRS Document",
         "Antigravity (Gemini)", "Documentation",
         "Prompted AI to generate full SRS structure (functional requirements per module, non-functional requirements, business rules, use case catalog) based on the existing codebase and database schema.",
         "Verified all FR-IDs trace to actual implemented controllers/services. Corrected estimated user counts and performance thresholds against test data. Added VIP slot, lost ticket, and subscription edge cases.",
         "docs/SRS_PBMS_SU26SWP08.md → .docx",
         "Accepted with modifications"),

        ("AI-018", "Doc Phase\n2026-07-20", "UML Diagrams\n(Use Case, Sequence, Class)",
         "Antigravity (Gemini)", "Documentation",
         "Used AI to generate 40+ PlantUML diagrams including overall use case, actor-specific use cases, check-in/checkout sequence diagrams, and class diagram.",
         "Reviewed actor associations and include/extend relationships. Corrected sequence diagram activation bars. Verified class diagram against EF Core model classes.",
         "docs/UML_Diagrams_PBMS_SU26SWP08.md → .docx",
         "Accepted with modifications"),

        ("AI-019", "Doc Phase\n2026-07-23", "SDS Document",
         "Antigravity (Gemini)", "Documentation",
         "Prompted AI to generate Software Design Specification covering architecture topology, technology stack, service responsibilities, API guidelines, SignalR group design, and deployment strategy.",
         "Verified all service descriptions against actual Service layer implementations. Confirmed endpoint paths match API routing attributes.",
         "docs/SDS_PBMS_SU26SWP08.md → .docx",
         "Accepted with modifications"),

        ("AI-020", "Doc Phase\n2026-07-23", "Testing Report",
         "Antigravity (Gemini)", "Documentation",
         "Prompted AI to draft Testing Report covering unit test cases (PricingService, TicketCodeGenerator, VietnamesePlateParser), integration tests (ParkingSessionFlowTests, ReservationFlowTests), E2E Playwright specs, and performance metrics.",
         "Cross-referenced all TC IDs against actual xUnit test files in ParkingBuildingManagement.Api.Tests. Corrected expected values to match actual test assertions.",
         "docs/Testing_Report_PBMS_SU26SWP08.md → .docx",
         "Accepted with modifications"),

        ("AI-021", "Doc Phase\n2026-07-23", "User Manual\n& Presentation",
         "Antigravity (Gemini)", "Documentation",
         "Prompted AI to generate User Manual (4 role sections: Admin, Manager, Staff, Driver) and 16-slide graduation presentation PPTX generation script.",
         "Verified all UI flow descriptions against actual frontend screens. Corrected payment method descriptions and OCR workflow steps.",
         "docs/User_Manual_PBMS_SU26SWP08.md, scripts/generate_pptx.py",
         "Accepted with modifications"),
    ]

    for i, row_data in enumerate(rows):
        alt = (i % 2 == 1)
        r = i + 3
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 70

    set_col_widths(ws, [10, 14, 22, 20, 18, 55, 48, 42, 22])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Code Generation Detail
# ═══════════════════════════════════════════════════════════════════════════════
def build_code_gen(wb):
    ws = wb.create_sheet("4. Code Generation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Code Generation — Detailed Breakdown by Component"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    # ── Backend
    r = 2
    section_title(ws, r, 1, "Backend — ASP.NET Core API", 7, bg="1A5276")
    r += 1
    headers = ["Component", "File(s)", "AI Tool", "AI-Generated Portion", "Developer Modification", "Review Effort", "Final Status"]
    apply_header_row(ws, r, headers, bg="2980B9")
    r += 1

    backend_rows = [
        ("AuthService.cs", "Services/AuthService.cs", "Cursor + GitHub Copilot",
         "JWT token generation helper, BCrypt hash call, Google token exchange scaffold",
         "Added role-claim injection, 24-hour expiry config, refresh-token stub placeholder removed",
         "High", "Production-ready"),
        ("ParkingSessionService.cs\n(check-in)", "Services/ParkingSessionService.cs", "Cursor (Chat)",
         "Serializable transaction wrapper, slot status update sequence, session entity initialisation",
         "Added plate normalisation, reservation linking logic, SlotPosition validation, NotifyCheckInAsync call",
         "High", "Production-ready"),
        ("ParkingSessionService.cs\n(check-out)", "Services/ParkingSessionService.cs", "Cursor (Chat) + GitHub Copilot",
         "Fee deduction from wallet, session status transition, payment record creation",
         "Added subscription waiver check, penalty + VIP surcharge aggregation, idempotency guard",
         "High", "Production-ready"),
        ("PricingService.cs", "Services/PricingService.cs", "GitHub Copilot",
         "Grace period check, ceil(hours) multiplication, daily cap comparison",
         "Corrected boundary condition for exact-grace-period case; added lost ticket fee branch; verified VND precision",
         "Medium", "Production-ready"),
        ("SlotAllocationService.cs", "Services/SlotAllocationService.cs", "Cursor (Inline)",
         "LINQ query for least-occupied zone, ascending SlotPosition ordering",
         "Added VIP slot bypass path (SlotPosition == 1 reserve); added explicit slot override support",
         "Medium", "Production-ready"),
        ("ReservationService.cs", "Services/ReservationService.cs", "ChatGPT → adapted",
         "Reservation lifecycle logic (Pending → Confirmed transition), expiry background service pattern",
         "Verified state machine correctness; implemented max-active-reservations cap from SystemConfig; wired SignalR notify",
         "High", "Production-ready"),
        ("WalletService.cs", "Services/WalletService.cs", "ChatGPT template adapted",
         "PayOS QR link generation call, webhook credit logic skeleton",
         "Added idempotency (OrderCode uniqueness check); added demo-mode top-up bypass; wired SignalR notification",
         "High", "Production-ready"),
        ("ParkingRealtimeNotifier.cs", "Services/ParkingRealtimeNotifier.cs", "Cursor (Chat)",
         "SignalR group send pattern for slotUpdated, dashboardRefresh, wallet events",
         "Verified group naming vs hub.cs; added driver-targeted group driver:{id} for wallet notifications",
         "Medium", "Production-ready"),
        ("API Controllers", "Controllers/*.cs (22 files)", "Cursor (Inline) + GitHub Copilot",
         "CRUD routing structure, [Authorize(Roles=...)] attribute placement, DTO binding patterns",
         "Added business exception mapping; verified role assignments against SRS; added pagination parameters",
         "High", "Production-ready"),
        ("PayOsWebhookController.cs", "Controllers/PayOsWebhookController.cs", "ChatGPT",
         "HMAC-SHA256 signature verification pattern for incoming webhook payload",
         "Adapted HMAC key sourcing from configuration; added logging; verified against PayOS documentation",
         "High", "Production-ready"),
        ("GoogleTokenValidator.cs", "Services/GoogleTokenValidator.cs", "Cursor (Chat)",
         "HTTP call to Google tokeninfo endpoint for ID token validation",
         "Replaced raw HTTP with Google .NET Auth Library; added failure fallback; secured against token replay",
         "High", "Production-ready"),
        ("OCR Services", "Services/Ocr/ (2 files)", "ChatGPT",
         "HttpClient call to local PaddleOCR server; fallback pattern to PlateRecognizer cloud API",
         "Added Vietnamese plate regex validation; implemented local → cloud fallback priority; stubbed Linux path",
         "High", "Production-ready"),
        ("xUnit Test Scaffolding", "ParkingBuildingManagement.Api.Tests/*.cs (13 files)", "GitHub Copilot + Cursor",
         "InMemory DbContext setup (PbmsWebApplicationFactory), test data seeding helpers (TestData.cs), xUnit theory attribute patterns",
         "Validated DI wiring; corrected InMemory vs real-PostgreSQL divergence in Serializable tests; wrote all assertion logic manually",
         "High", "Production-ready"),
        ("DatabaseSeeder.cs", "Services/DatabaseSeeder.cs", "GitHub Copilot",
         "Initial seed data loop for vehicle types, zones, slots scaffold",
         "Expanded to seed 104 zones × 8 slots, 5 vehicle types, pricing policies, subscription plans, admin/manager/staff accounts",
         "Medium", "Production-ready"),
    ]

    for i, row_data in enumerate(backend_rows):
        alt = (i % 2 == 1)
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 58
        r += 1

    # ── Frontend
    r += 1
    section_title(ws, r, 1, "Frontend — React 18 + Vite + Tailwind CSS", 7, bg="117A65")
    r += 1
    apply_header_row(ws, r, headers, bg="1ABC9C")
    r += 1

    frontend_rows = [
        ("React Router\nNested Routes", "src/app/router.tsx, src/app/layouts/", "Cursor (Chat)",
         "Nested route hierarchy scaffold for /admin, /manager, /staff, /driver with ProtectedRoute HOC",
         "Added token-expiry redirect; integrated role validation from JWT claims; corrected lazy-loaded page imports",
         "Medium", "Production-ready"),
        ("Axios Interceptor\n& Auth Context", "src/app/hooks/useAuth.ts, src/app/api/axiosClient.ts", "Cursor (Inline) + GitHub Copilot",
         "Axios instance with Bearer token injection; 401 response interceptor for auto-logout",
         "Added token refresh placeholder; integrated with React Context; verified cross-tab logout behaviour",
         "Medium", "Production-ready"),
        ("SignalR Hook", "src/app/hooks/useSignalR.ts", "Cursor (Chat)",
         "useEffect-based SignalR connection lifecycle with reconnect policy",
         "Fixed memory leak (dispose on unmount); integrated with slot map state; handled driver-group subscriptions",
         "High", "Production-ready"),
        ("Slot Map Visualiser", "src/app/pages/SlotMap/", "Cursor (Chat) + Cursor (Inline)",
         "Grid rendering per floor/zone, colour-coded status cells (Available/Occupied/Reserved/Maintenance)",
         "Added click-to-checkout action; integrated live SignalR updates; optimised re-render with React.memo",
         "High", "Production-ready"),
        ("Check-In Form", "src/app/pages/CheckIn/", "Cursor (Chat)",
         "Multi-step form scaffold (plate input → vehicle type select → zone select → confirm)",
         "Added OCR scan button integration; linked reservation ID field; added slot override toggle",
         "Medium", "Production-ready"),
        ("Check-Out Flow", "src/app/pages/CheckOut/", "Cursor (Inline)",
         "Fee preview component, payment method radio group, confirm button",
         "Added lost-ticket toggle; integrated subscription badge display; handled EWallet insufficient funds UX",
         "Medium", "Production-ready"),
        ("Wallet Top-Up\n(PayOS + Demo Mode)", "src/app/pages/Wallet/", "ChatGPT template → adapted",
         "PayOS QR code display component, polling for payment status",
         "Added demo-mode simulate button; added balance refresh on webhook success; improved error state handling",
         "Medium", "Production-ready"),
        ("Manager Dashboard\n& Charts", "src/app/pages/Dashboard/", "Cursor (Chat) + GitHub Copilot",
         "KPI card grid scaffold, chart library integration (Recharts) for 30-day revenue trend",
         "Verified KPI data mapping against API response shape; added occupancy % ring chart; responsive layout fixes",
         "Medium", "Production-ready"),
        ("Google Sign-In\nButton", "src/app/pages/Auth/GoogleLoginButton.tsx", "Cursor (Chat)",
         "Google Identity Services script loader, credential response handler",
         "Added error handling for token validation failure; restricted display to Driver login page only",
         "Low", "Production-ready"),
        ("E2E Playwright Tests", "frontend/e2e/smoke.spec.ts", "GitHub Copilot",
         "Basic smoke test spec structure (login, navigate, assert)",
         "Added correct DOM selectors matching actual rendered HTML; expanded to cover staff check-in flow",
         "Medium", "Production-ready"),
    ]

    for i, row_data in enumerate(frontend_rows):
        alt = (i % 2 == 1)
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 55
        r += 1

    set_col_widths(ws, [28, 38, 25, 55, 52, 16, 18])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Documentation Assistance
# ═══════════════════════════════════════════════════════════════════════════════
def build_docs(wb):
    ws = wb.create_sheet("5. Documentation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Documentation Assistance — AI Usage in Technical Documents"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Document", "AI Tool(s)", "Sections AI-Assisted", "AI Drafting Contribution", "Developer Review Actions", "Review Completeness", "Final Status"]
    apply_header_row(ws, 2, headers)

    docs = [
        ("SRS\n(Software Requirements Specification)",
         "Antigravity / Gemini + ChatGPT",
         "All sections: FR catalog, NFRs, Business Rules, Use Case Catalog, Traceability Matrix, Appendix",
         "~60% initial draft",
         "Verified all FR-IDs trace to actual service/controller code. Corrected user count estimates. Added VIP, lost-ticket, grace-period edge cases. Reviewed all 42 use cases against actual implementation.",
         "100%", "Final — Approved"),
        ("SDS\n(Software Design Specification)",
         "Antigravity / Gemini",
         "Architecture topology diagram, technology stack table, service responsibility descriptions, API endpoint list, SignalR group design, deployment pipeline",
         "~55% initial draft",
         "Verified service descriptions against source files. Corrected endpoint paths to match controller Route attributes. Confirmed SignalR hub group names vs ParkingRealtimeNotifier.cs.",
         "100%", "Final — Approved"),
        ("UML Diagrams",
         "Antigravity / Gemini",
         "Overall use case diagram (42 UCs), actor-specific detailed diagrams (Admin/Manager/Staff/Driver), check-in sequence diagram, check-out sequence diagram, reservation lifecycle diagram, class diagram",
         "~70% initial PlantUML draft",
         "Reviewed all actor-UC associations and include/extend relationships. Corrected sequence diagram activation lifelines. Verified class diagram against 17+ EF Core model classes and navigation properties.",
         "100%", "Final — Approved"),
        ("Testing Report",
         "Antigravity / Gemini",
         "Test methodology, all TC tables (unit/integration/E2E), performance profile table, defect summary",
         "~65% initial draft",
         "Cross-referenced TC-PRC-001 to TC-PRC-004 against PricingService unit tests. Verified TC-INT-001 to TC-INT-004 against ParkingSessionFlowTests.cs and ReservationFlowTests.cs. Confirmed P50/P95 timing against actual profiling runs.",
         "100%", "Final — Approved"),
        ("User Manual",
         "Antigravity / Gemini",
         "All 4 role sections (Admin, Manager, Staff, Driver), step-by-step workflows, FAQ",
         "~50% initial draft",
         "Verified all UI workflow descriptions against actual frontend screens. Corrected payment flow for demo-mode vs live PayOS. Added OCR troubleshooting steps based on actual system behaviour.",
         "100%", "Final — Approved"),
        ("Issues Report",
         "Antigravity / Gemini + ChatGPT",
         "Issue log table structure, classification taxonomy",
         "~30% structure draft",
         "All 15+ issue entries written manually based on actual development incidents. AI only provided table structure template.",
         "100%", "Final — Approved"),
        ("Presentation Slides\n(PPTX — 16 slides)",
         "Antigravity / Gemini",
         "Slide content outline, speaker notes, generate_pptx.py script for automated deck generation",
         "~45% outline + script",
         "Verified all statistics (API endpoints count, test count, sprint timeline). Adapted slide layouts to 45-min defence format. Generated final PPTX using python-pptx script.",
         "100%", "Final — Approved"),
        ("README.md\n(GitHub)",
         "ChatGPT",
         "Project overview, technology badges, quick-start commands, architecture section",
         "~40% initial draft",
         "Added deployment-specific instructions (Railway + Firebase). Corrected environment variable names to match actual appsettings.example.json.",
         "100%", "Final — Approved"),
        ("deploy-cicd.md\n& security runbooks",
         "ChatGPT",
         "CI/CD pipeline explanation, Gitleaks setup guide, secret rotation runbook",
         "~50% initial draft",
         "Verified all commands against actual workflow YAML. Added Railway-specific deployment steps. Corrected Firebase hosting target name.",
         "100%", "Final — Approved"),
    ]

    for i, row_data in enumerate(docs):
        alt = (i % 2 == 1)
        apply_body_row(ws, 3 + i, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[3 + i].height = 65

    set_col_widths(ws, [28, 24, 42, 20, 55, 18, 18])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – Metrics & Compliance
# ═══════════════════════════════════════════════════════════════════════════════
def build_metrics(wb):
    ws = wb.create_sheet("6. Metrics & Compliance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Usage Metrics & Compliance Controls"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    # ── AI Drafting Ratio Table
    r = 3
    section_title(ws, r, 1, "AI Drafting Ratio by Project Area", 6)
    r += 1
    apply_header_row(ws, r, ["Project Area", "Component Count", "AI Draft Ratio (%)", "Human Review Ratio (%)", "Accepted As-Is (%)", "Accepted w/ Modification (%)"])
    r += 1

    area_rows = [
        ("Authentication & Security", "6 files", "35%", "100%", "0%", "100%"),
        ("Core Parking Operations (Check-In/Out)", "8 files", "30%", "100%", "0%", "100%"),
        ("Payment & Wallet (PayOS Integration)", "5 files", "40%", "100%", "0%", "100%"),
        ("Reservation Management", "4 files", "35%", "100%", "0%", "100%"),
        ("Subscription System", "5 files", "40%", "100%", "0%", "100%"),
        ("SignalR Real-Time Notifications", "3 files", "45%", "100%", "0%", "100%"),
        ("OCR / Plate Recognition", "4 files", "50%", "100%", "0%", "100%"),
        ("Reporting & Analytics", "3 files", "40%", "100%", "0%", "100%"),
        ("API Controllers (CRUD Routing)", "22 files", "45%", "100%", "0%", "100%"),
        ("xUnit / Integration Tests", "13 files", "25%", "100%", "0%", "100%"),
        ("Frontend – React Components", "~40 files", "35%", "100%", "0%", "100%"),
        ("Frontend – E2E Playwright Tests", "1 file", "30%", "100%", "0%", "100%"),
        ("Infrastructure / CI-CD / Docker", "8 files", "40%", "100%", "0%", "100%"),
        ("Technical Documentation (MD/DOCX)", "9 documents", "55%", "100%", "0%", "100%"),
    ]

    for i, row_data in enumerate(area_rows):
        alt = (i % 2 == 1)
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Sprint Timeline
    r += 2
    section_title(ws, r, 1, "AI Usage by Sprint / Month", 6, bg="6C3483")
    r += 1
    apply_header_row(ws, r, ["Sprint", "Period", "Primary AI Tools", "Key AI-Assisted Tasks", "Sessions Estimate", "Notes"], bg="8E44AD")
    r += 1

    sprint_rows = [
        ("Sprint 1", "May 30 – Jun 16, 2026", "ChatGPT, GitHub Copilot, Cursor", "Project scaffold, CI/CD workflows, initial DB models, JWT auth skeleton", "~15 sessions", "Foundation sprint; heavy consultation phase"),
        ("Sprint 2", "Jun 16 – Jun 20, 2026", "Cursor (Chat+Inline), GitHub Copilot", "PayOS webhook, Role portals (React), Rate limiting, Security headers", "~20 sessions", "Integration sprint; most AI code review sessions"),
        ("Sprint 3", "Jun 20 – Jun 26, 2026", "Cursor (Chat+Inline), ChatGPT", "Check-in/out Serializable flow, SignalR hub, Slot map visualiser, OCR scanner", "~25 sessions", "Core business logic sprint; high AI assistance"),
        ("Sprint 4", "Jun 26 – Jul 14, 2026", "ChatGPT, Cursor, GitHub Copilot", "Reservation lifecycle, Google OAuth, VIP slots, Security hardening, Subscriptions", "~20 sessions", "Feature-complete sprint; complex integrations"),
        ("Sprint Final", "Jul 14 – Jul 24, 2026", "Antigravity/Gemini, ChatGPT", "All documentation generation, UML diagrams, Testing Report, User Manual, Presentation", "~30 sessions", "Documentation-focused; highest doc AI usage"),
    ]

    for i, row_data in enumerate(sprint_rows):
        alt = (i % 2 == 1)
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 40
        r += 1

    # ── Compliance Controls
    r += 2
    section_title(ws, r, 1, "Security & Compliance Controls", 6, bg="1A5276")
    r += 1
    apply_header_row(ws, r, ["Control ID", "Control Rule", "Zone / Scope", "Implementation", "Verified By", "Status"], bg="2980B9")
    r += 1

    controls = [
        ("CC-001", "No production secrets in AI prompts", "All AI tools", "Connection strings, JWT signing keys, PayOS API keys, Google Client Secret never included in any prompt. Environment variables used exclusively.", "Peer code review + .gitleaks.toml scan", "Compliant"),
        ("CC-002", "AI output review before integration", "All code & docs", "Every AI-generated code block reviewed by developer before commit. Docs cross-referenced against source code before export.", "Git commit history + PR review", "Compliant"),
        ("CC-003", "Hallucination identification protocol", "All AI tools", "AI outputs that introduced non-existent APIs, incorrect EF Core patterns, or wrong PayOS endpoint signatures were identified and discarded during peer review.", "Developer review log (AI-004 to AI-016)", "Compliant"),
        ("CC-004", "No personal/customer data in prompts", "All AI tools", "No real license plate numbers, user emails, wallet transaction data, or PII included in any AI prompt. Anonymised examples used for context only.", "Team agreement + session logs", "Compliant"),
        ("CC-005", "AI usage transparency disclosure", "Project documentation", "This AI Usage Report documents all tool interactions per SWP391 academic requirements.", "This report", "Compliant"),
        ("CC-006", "Cursor pre-push security rules", "Cursor IDE", "cursor-pre-push-security-rules.mdc enforces secret pattern scan before any git push via Cursor.", ".cursor/rules/, .gitleaksignore", "Compliant"),
    ]

    for i, row_data in enumerate(controls):
        alt = (i % 2 == 1)
        apply_body_row(ws, r, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[r].height = 48
        r += 1

    set_col_widths(ws, [14, 42, 25, 55, 28, 14])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – Lessons Learned
# ═══════════════════════════════════════════════════════════════════════════════
def build_lessons(wb):
    ws = wb.create_sheet("7. Lessons Learned")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Lessons Learned — AI Integration in Software Development"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    ws["A1"].fill = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["Theme", "Observation", "Impact", "Recommendation", "Rating"]
    apply_header_row(ws, 2, headers)

    lessons = [
        ("Code Quality",
         "Cursor inline completions significantly accelerated CRUD controller and DTO generation, reducing per-controller setup time from ~45 min to ~15 min. However, Copilot occasionally suggested outdated EF Core API patterns (DbContext.Entry vs direct property access).",
         "Positive: Faster scaffold generation.\nRisk: Requires developer to recognise outdated API suggestions.",
         "Always verify AI-generated ORM code against current framework documentation. Run unit tests immediately after integrating AI suggestions.",
         "★★★★☆"),
        ("Architecture Design",
         "ChatGPT was highly effective for discussing architectural trade-offs (e.g., Serializable vs Repeatable Read isolation for the check-in flow). Provided clear explanations and code examples.",
         "Positive: Improved design confidence for critical financial operations.",
         "Use AI for architectural consultation early in planning phase rather than as a code generator for complex business logic.",
         "★★★★★"),
        ("Security",
         "AI correctly identified HMAC verification as the required pattern for PayOS webhooks and rate-limiting as essential for auth endpoints. However, AI did not proactively suggest the Gitleaks integration — this required developer initiative.",
         "Positive: Security patterns suggested accurately.\nGap: AI does not proactively audit for all security requirements.",
         "Maintain a developer-owned security checklist. Do not rely solely on AI to surface security gaps.",
         "★★★★☆"),
        ("Documentation",
         "Antigravity (Gemini) produced high-quality structured documentation drafts (SRS, SDS, UML, User Manual) that matched academic standards. Reduced documentation time by approximately 60%. PlantUML diagram generation was particularly effective.",
         "Positive: Major time saving on documentation.\nRisk: AI may include plausible-sounding but inaccurate feature descriptions that do not match actual implementation.",
         "Always cross-reference AI-generated documentation paragraph by paragraph against actual source code. Use the codebase as ground truth.",
         "★★★★★"),
        ("Testing",
         "GitHub Copilot effectively scaffolded xUnit test class boilerplate and InMemory provider setup. However, test assertions and edge case data were entirely developer-written — AI suggestions were too generic.",
         "Positive: Faster test setup.\nLimitation: Test quality depends on developer-written assertions.",
         "Use AI only for test harness setup (factories, DI wiring). Write all assertions and test data manually.",
         "★★★☆☆"),
        ("Hallucination Risk",
         "Across ~110 AI sessions, approximately 5–8 instances of clear hallucination were identified: non-existent .NET 10 APIs, incorrect PayOS webhook field names, wrong SignalR method signatures. All were caught during developer review.",
         "Risk: Hallucinated code can introduce subtle bugs if not reviewed carefully.",
         "Treat AI output as a first draft requiring mandatory compilation + test verification. Never push AI-generated code directly without review.",
         "★★★☆☆"),
        ("Productivity",
         "Overall development velocity increased by an estimated 25–35% compared to a non-AI baseline. Biggest gains were in boilerplate generation (controllers, DTOs, migrations) and documentation drafting.",
         "Positive: Meaningful productivity improvement.\nNote: Productivity gains concentrated in mechanical tasks, not complex algorithm design.",
         "Reserve AI assistance for high-volume repetitive tasks. invest saved time in deeper testing and architectural review.",
         "★★★★★"),
    ]

    for i, row_data in enumerate(lessons):
        alt = (i % 2 == 1)
        apply_body_row(ws, 3 + i, row_data, alt=alt, bold_first=True)
        ws.row_dimensions[3 + i].height = 72

    set_col_widths(ws, [22, 55, 40, 50, 12])


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    build_cover(wb)
    build_tools(wb)
    build_usage_log(wb)
    build_code_gen(wb)
    build_docs(wb)
    build_metrics(wb)
    build_lessons(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "docs", "Exported_Plain")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "AI_Usage_Report_PBMS_SU26SWP08.xlsx")
    wb.save(out_path)
    print(f"[OK] Generated: {os.path.abspath(out_path)}")


if __name__ == "__main__":
    main()
